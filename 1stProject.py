import pandas as pd
import numpy as np
from scipy.stats import boxcox
import matplotlib.pyplot as plt
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import OneHotEncoder
from sklearn.preprocessing import OrdinalEncoder
from sklearn.preprocessing import MinMaxScaler
import seaborn as sns
from statsmodels.api import add_constant, Logit
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, precision_score, recall_score, classification_report, accuracy_score
from imblearn.over_sampling import SMOTE
from sklearn.metrics import accuracy_score
from sklearn.metrics import classification_report
from openpyxl import load_workbook
from imblearn.combine import SMOTEENN
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import roc_auc_score, roc_curve
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import (
    classification_report,
    confusion_matrix,
    roc_auc_score,
    precision_recall_curve,
    RocCurveDisplay)
from sklearn.model_selection import StratifiedKFold
from imblearn.combine import SMOTEENN
from sklearn.metrics import roc_auc_score, precision_recall_curve, RocCurveDisplay

#load data
dataFrame = pd.read_csv("C:/Users/user/Downloads/transactions_wholesale_with_issues(1).csv")
print(dataFrame.head())

#********************************************************************************************>
# Rename columns using Camel Style
print(dataFrame.columns)

dataFrame.columns = [
    ''.join(word.capitalize() if i != 0 else word.lower() for i, word in enumerate(col.replace('-', ' ').replace('_', ' ').split()))
    for col in dataFrame.columns]

print(dataFrame.columns)

#print data info
dataFrame.info()
#print data description
print(dataFrame.describe(include='all'))

#********************************************************************************************>
# We calculate and plot correlartion matrix
# Extract numeric columns
numeric_col = dataFrame.select_dtypes(exclude=['object']).columns

# Calculate the correlation matrix, ignoring null values
correlation_matrix = dataFrame[numeric_col].corr(method='pearson', min_periods=1)

# Plot the heatmap
plt.figure(figsize=(10, 8))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)

# Add title and labels
plt.title("Correlation Matrix Heatmap for Numeric Features (Nulls Ignored), using Original Data Set", fontsize=16)
plt.xlabel("Features", fontsize=12)
plt.ylabel("Features", fontsize=12)

# Display the heatmap
plt.tight_layout()  # Adjust layout to fit everything
plt.show()
#********************************************************************************************<

#********************************************************************************************>
# Now we calculate the number/percentage of missing values for each variable in our data set
print("******************************************************************************")
print("Missing values' Frequency Distribution through or data set \n")
print(dataFrame.isnull().sum(), '\n')
print( round( dataFrame.isnull().mean() *100, 2))

# Now we calculate the total number/percentage of rows with at least one missing value
rows_with_missing = dataFrame.isnull().any(axis = 1).sum()
row_withmissing_percentage = round(dataFrame.isnull().any(axis = 1).mean() *100, 2)

print("\n")
if(row_withmissing_percentage < 10 ):
    print(f"Total number of rows with at least one missing value {rows_with_missing}, which represents {row_withmissing_percentage} % out of our data set, which is less than 10%, so we decided to delete missing cases \n")
elif(row_withmissing_percentage > 10 ):
    print(f"Total number of rows with at least one missing value {rows_with_missing}, which represents {row_withmissing_percentage} % out of our data set, which is more than 10%, so we decided to impute missing cases \n")

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>*******************************<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
#-------------------------------------------------------------------------------------------------------------------------------
# Now we start by processing our continuous data variables

# We display the distribution of variabls, and determine the appropriate imputation method
# We plot the histogram for each quantitave data column using to its distribution
skewness_before_imputation = dataFrame.select_dtypes(exclude=['object']).skew()

for col in dataFrame.select_dtypes(exclude=['object']).columns:
    print(f"We create a histogram for {col} using Pandas \n")
    plt.figure(figsize=(8, 6))  
    dataFrame[col].plot(kind='hist', bins=30, color='green', edgecolor='black', alpha=0.7)
    plt.title(f"Histogram of {col}")
    plt.xlabel(col)
    plt.ylabel("Frequency")
    plt.grid(True)
    plt.show()
    print("\n")
    
    skewness = dataFrame[col].skew()
    
    print(f"Skewness value of the distribution of {col} is", skewness)
    if -1 <= skewness <= 1:
        print(f"The distribution of {col} is approximately symmetric (between -1 and 1). Therefor, we decided to impute using mean. \n")
        dataFrame[col].fillna(dataFrame[col].mean(), inplace=True)
    
    elif skewness > 1:
        print(f"The distribution of {col} is approximately psitively skewed (more than 1). Therefor, we decided to impute using median. \n")
        dataFrame[col].fillna(dataFrame[col].median(), inplace=True)
    
    elif skewness < -1:
        print(f"The distribution of {col} is approximately negatively skewed (less than -1). Therefor, we decided to impute using median. \n")
        dataFrame[col].fillna(dataFrame[col].median(), inplace=True)

#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Now after imputation, we calculate the number/percentage of missing values for each variable in our data set
print("******************************************************************************")
print("Missing values' Frequency Distribution through or data set \n")
print(dataFrame.isnull().sum(), '\n')
print( round( dataFrame.isnull().mean() *100, 2))

# Now after imputation, we calculate the total number/percentage of rows with at least one missing value
rows_with_missing = dataFrame.isnull().any(axis = 1).sum()
row_withmissing_percentage = round(dataFrame.isnull().any(axis = 1).mean() *100, 2)

print(f"Total number of rows with at least one missing value {rows_with_missing}, which represents {row_withmissing_percentage} % out of our data set.\n")

# Now after imputation, we plot the histogram for each quantitave data column using to its distribution
skewness_after_imputation = dataFrame.select_dtypes(exclude=['object']).skew()

for col in dataFrame.select_dtypes(exclude=['object']).columns:
    print(f"We create a histogram for {col} using Pandas \n")
    plt.figure(figsize=(8, 6))  
    dataFrame[col].plot(kind='hist', bins=30, color='green', edgecolor='black', alpha=0.7)
    plt.title(f"Histogram of {col}")
    plt.xlabel(col)
    plt.ylabel("Frequency")
    plt.grid(True)
    plt.show()
    print("\n")
    
    skewness = dataFrame[col].skew()
    
    print(f"Skewness value of the distribution of {col} is", skewness)
    if -1 <= skewness <= 1:
        print(f"The distribution of {col} is approximately symmetric (between -1 and 1). Therefor, we decided to impute using mean. \n")
        #dataFrame[col].fillna(dataFrame[col].mean(), inplace=True)
    
    elif skewness > 1:
        print(f"The distribution of {col} is approximately psitively skewed (more than 1). Therefor, we decided to impute using median. \n")
        #dataFrame[col].fillna(dataFrame[col].median(), inplace=True)
    
    elif skewness < -1:
        print(f"The distribution of {col} is approximately negatively skewed (less than -1). Therefor, we decided to impute using median. \n")
        #dataFrame[col].fillna(dataFrame[col].median(), inplace=True)
# Print skewness before and after imputation
print("Skewness before imputation:\n", skewness_before_imputation, "\n")
print("Skewness after imputation:\n", skewness_after_imputation)

#*********************************************************************************************************************<

#********************************************************************************************>
# We calculate and plot correlartion matrix
# Extract numeric columns
numeric_col = dataFrame.select_dtypes(exclude=['object']).columns

# Calculate the correlation matrix, ignoring null values
correlation_matrix = dataFrame[numeric_col].corr(method='pearson', min_periods=1)

# Plot the heatmap
plt.figure(figsize=(10, 8))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)

# Add title and labels
plt.title("Correlation Matrix Heatmap for Numeric Features, using Imputed Data Set", fontsize=16)
plt.xlabel("Features", fontsize=12)
plt.ylabel("Features", fontsize=12)

# Display the heatmap
plt.tight_layout()  # Adjust layout to fit everything
plt.show()


# Now We handle the skewed distributions using the suitable mathematical transformation based on the value of skewness
skewness_values = dataFrame.select_dtypes(exclude=['object']).skew()  # Calculate skewness for all numeric columns

for col in dataFrame.select_dtypes(exclude=['object']).columns:
    if col == 'class':  # Exclude the 'class' column
        print(f"Skipping column '{col}' as it is a categorical variable, the dependent variable.")
        continue  

    skewness = skewness_values[col]
    print(f"Processing column '{col}' with skewness: {skewness:.4f}")

    if skewness >= 5:  # Extremely positively skewed
        print(f"Column '{col}' has extreme skewness (>6). Applying Box-Cox transformation.")
        dataFrame[col], _ = boxcox(dataFrame[col] + 1)  # Add 1 to avoid zero values
    
    elif 3 < skewness < 5:  # Highly positively skewed
        print(f"Column '{col}' is highly positively skewed (>3). Applying log transformation.")
        dataFrame[col] = np.log1p(dataFrame[col])  # log(1 + x)
    
    elif 1 < skewness <= 3:  # Moderately positively skewed
        print(f"Column '{col}' is moderately positively skewed (>1). Applying square root transformation.")
        dataFrame[col] = np.sqrt(dataFrame[col])
    
    elif -1 < skewness <= 1:  # Approximately symmetric
        print(f"Column '{col}' is approximately symmetric. No transformation applied.")
    
    elif -3 <= skewness < -1:  # Highly negatively skewed
        print(f"Column '{col}' is highly negatively skewed (<-1). Applying reflection + log transformation.")
        max_value = dataFrame[col].max()
        dataFrame[col] = np.log1p(max_value - dataFrame[col])  # log(1 + max_value - x)
    
    elif skewness < -3:  # Extremely negatively skewed
        print(f"Column '{col}' has extreme negative skewness (<-3). Applying reflection + Box-Cox transformation.")
        max_value = dataFrame[col].max()
        dataFrame[col], _ = boxcox(max_value - dataFrame[col] + 1)  # Box-Cox for reflection
    
    else:  # If skewness is very close to 0 (symmetric)
        print(f"Column '{col}' is approximately symmetric or no transformation needed.")

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Now after the mathematical transformation, we plot the histogram for each quantitave data column using to its distribution
skewness_after_imputation = dataFrame.select_dtypes(exclude=['object']).skew()

for col in dataFrame.select_dtypes(exclude=['object']).columns:
    print(f"We create a histogram for {col} using Pandas \n")
    plt.figure(figsize=(8, 6))  
    dataFrame[col].plot(kind='hist', bins=30, color='green', edgecolor='black', alpha=0.7)
    plt.title(f"Histogram of {col}")
    plt.xlabel(col)
    plt.ylabel("Frequency")
    plt.grid(True)
    plt.show()
    print("\n")

Skewness_after_Mathematical_transformation = dataFrame.select_dtypes(exclude=['object']).skew()
# Now we print skewness skewness before and after imputation, and after the mathematical transformation
print("Skewness before imputation:\n", skewness_before_imputation, "\n")
print("Skewness after imputation:\n", skewness_after_imputation, "\n")
print("Skewness after transformation:\n", Skewness_after_Mathematical_transformation, "\n")
#********************************************************************************************>

#********************************************************************************************>
# We calculate and plot correlartion matrix
# Extract numeric columns
numeric_col = dataFrame.select_dtypes(exclude=['object']).columns

# Calculate the correlation matrix, ignoring null values
correlation_matrix = dataFrame[numeric_col].corr(method='pearson', min_periods=1)

# Plot the heatmap
plt.figure(figsize=(10, 8))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)

# Add title and labels
plt.title("Correlation Matrix Heatmap for Numeric Features, using Imputed and Transformed Data Set", fontsize=16)
plt.xlabel("Features", fontsize=12)
plt.ylabel("Features", fontsize=12)

# Display the heatmap
plt.tight_layout()  # Adjust layout to fit everything
plt.show()
#********************************************************************************************<
#-------------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------------
#********************************************************************************************<
# Now we unify the scales for continuos variables, using MinMax Scaler
print(dataFrame.describe(include='all'),'\n')
# Select numerical columns and exclude 'risk'
quantDataToScale = dataFrame.select_dtypes(exclude=['object']).drop(columns=['class']).columns
print(quantDataToScale)
# Initialize MinMaxScaler
scaler = MinMaxScaler()
# Fit and transform the data
normalized_data = scaler.fit_transform(dataFrame[quantDataToScale])
# Create a DataFrame from the normalized data with "_normalized" suffix for the column names
dataFrame[quantDataToScale] = pd.DataFrame(normalized_data, columns=[f"{col}_normalized" for col in quantDataToScale])
# # Concatenate the normalized columns with the original DataFrame
# dataFrame = pd.concat([dataFrame, normalized_df], axis=1)

print(dataFrame,'\n')
print(dataFrame.describe(include='all'),'\n')
#********************************************************************************************<

#********************************************************************************************>
# We calculate and plot correlartion matrix
# Extract numeric columns
numeric_col = dataFrame.select_dtypes(exclude=['object']).columns

# Calculate the correlation matrix, ignoring null values
correlation_matrix = dataFrame[numeric_col].corr(method='pearson', min_periods=1)

# Plot the heatmap
plt.figure(figsize=(10, 8))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)

# Add title and labels
plt.title("Correlation Matrix Heatmap for Numeric Features, using Imputed ,Transformed, and Scaled Data Set", fontsize=16)
plt.xlabel("Features", fontsize=12)
plt.ylabel("Features", fontsize=12)

# Display the heatmap
plt.tight_layout()  # Adjust layout to fit everything
plt.show()
#********************************************************************************************<

#********************************************************************************************>
# Now we Continuo by processing our categorical data variables
#----------------------------------------------------------------------------------------------------
# We Check the frequency table for Categorical variable (Columns of type Object)
for col in dataFrame.select_dtypes(include='object').columns:
    print(f"The distribution of column '{col}' is as follows:\n")
    print(dataFrame[col].value_counts(), "\n")

# We correct typos in the 'Product Category' column
dataFrame['productCategory'].replace('Electonics', 'Electronics', inplace=True)

# We correct typos in the 'Payment Method' column
dataFrame['paymentMethod'].replace('Crdt Crd', 'Credit Card', inplace=True)

# We verify the corrections
print(dataFrame['productCategory'].value_counts())
print(dataFrame['paymentMethod'].value_counts())
#-----------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------
# We create bar plot for each categorical variable in our data
for col in dataFrame.select_dtypes(exclude=['number']).columns:
    print(f"We create a bar plot for '{col}' including missing values \n")
    # We create a copy of the column with 'Missing' as a placeholder for NaN values
    col_with_missing = dataFrame[col].fillna('Missing')
    # Calculate value counts and percentages
    value_counts = col_with_missing.value_counts()
    percentages = (value_counts / len(dataFrame[col])) * 100
    # Create a figure
    plt.figure(figsize=(8, 6))
    # Bar plot for categorical data with missing values included
    ax = value_counts.plot(kind='bar', color='blue', edgecolor='black', alpha=0.7)
    # Annotate bars with percentages
    for i, value in enumerate(value_counts):
        percent = percentages[i]
        ax.text(i, value + 0.5, f'{percent:.1f}%', ha='center', fontsize=10, color='black')
    # Customize plot
    plt.title(f"Bar Plot of {col} (Including Missing Values)")
    plt.xlabel(col)
    plt.ylabel("Count")
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
    plt.tight_layout()  # Ensure everything fits in the plot area
    # Display plot
    plt.show()
    print("\n")
    # Print value counts and percentages
    print(f"The distribution of column '{col}' is as follows:\n")
    print(value_counts, "\n")
    print("Percentages:\n", percentages.round(1), "\n")
#-----------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------
# Now we impute the missing data in the categorical (Ordinal and Nomina) variables.

# Define categorical variables
nominal_columns = ['paymentMethod', 'productCategory']  # Nominal variables
ordinal_column = 'customerAgeGroup'  # Ordinal variable

# We impute nominal variables (paymentMethod, productCategory), using Mode imputation as it is appropriate for nominal variables as they have no meaningful order.
for col in nominal_columns:
    # Replace 'Missing' or NaN with mode
    mode_value = dataFrame[col].mode()[0]
    dataFrame[col] = dataFrame[col].fillna(mode_value).replace('Missing', mode_value)
    print(f"Imputed missing values in '{col}' with mode: {mode_value}\n")

# We impute ordinal variable (customerAgeGroup), using Mode as it is used for simplicity since it equals the median for this variable (Young Adult) group.
# We treat "Unknown" cases as missing data since their percentage is only 3.0% (Less than 5%). With such a low proportion, the information they contribute is minimal.
ordinal_mapping = {'Teen': 1, 'Young Adult': 2, 'Adult': 3, 'Senior': 4, 'Missing': None, 'Unknown': None}
reverse_mapping = {v: k for k, v in ordinal_mapping.items()}  # To map back after imputation
# Map ordinal categories to numeric values
dataFrame['customerAgeGroup_mapped'] = dataFrame['customerAgeGroup'].map(ordinal_mapping)
# Impute missing values with the mode (mapped numeric value)
mode_value = dataFrame['customerAgeGroup_mapped'].mode()[0]
dataFrame['customerAgeGroup_mapped'] = dataFrame['customerAgeGroup_mapped'].fillna(mode_value)
# Map numeric values back to original categories
dataFrame['customerAgeGroup'] = dataFrame['customerAgeGroup_mapped'].map(reverse_mapping)
# Drop the temporary numeric column
dataFrame.drop(columns=['customerAgeGroup_mapped'], inplace=True)
#-----------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------
# Now after categorical variable imputation, wwe create bar plot for each categorical variable in our data
for col in dataFrame.select_dtypes(exclude=['number']).columns:
    print(f"We create a bar plot for '{col}' , after Mode Imputation \n")
    # We create a copy of the column with 'Missing' as a placeholder for NaN values
    col_with_missing = dataFrame[col].fillna('Missing')
    # Calculate value counts and percentages
    value_counts = col_with_missing.value_counts()
    percentages = (value_counts / len(dataFrame[col])) * 100
    # Create a figure
    plt.figure(figsize=(8, 6))
    # Bar plot for categorical data with missing values included
    ax = value_counts.plot(kind='bar', color='blue', edgecolor='black', alpha=0.7)
    # Annotate bars with percentages
    for i, value in enumerate(value_counts):
        percent = percentages[i]
        ax.text(i, value + 0.5, f'{percent:.1f}%', ha='center', fontsize=10, color='black')
    # Customize plot
    plt.title(f"Bar Plot of {col} (After Mode Imputation)")
    plt.xlabel(col)
    plt.ylabel("Count")
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
    plt.tight_layout()  # Ensure everything fits in the plot area
    # Display plot
    plt.show()
    print("\n")
    # Print value counts and percentages
    print(f"The distribution of column '{col}' is as follows:\n")
    print(value_counts, "\n")
    print("Percentages:\n", percentages.round(1), "\n")
#********************************************************************************************>
#----------------------------------------------------------------------------------------------------------------
# Now we classify categorical variables into Nominal and Ordinal
nominalColumns = ['productCategory','paymentMethod']
ordinalColumns = ['customerAgeGroup']
categoricalColumns = ['productCategory','paymentMethod','customerAgeGroup']

# 1) We use Lable Encoding for categorical variables, both Nominal and ordinal.
for col in dataFrame.select_dtypes(include='object').columns:
    encoder = LabelEncoder()
    dataFrame[f"{col}_Numeric_Nominal_Encoding"] = encoder.fit_transform(dataFrame[col]) + 1  # Add 1 to start from 1
print(dataFrame.columns)
print(dataFrame)
#////////////////////////////////////////////////////////////////////////////////////////////

# 2) We use One-Hot Encoding for categorical variables that are nominal (no inherent order).
#    The 'customerAgeGroup' variable is included in case it is treated as nominal in the future, disregarding the order of its levels.
# Instantiate the OneHotEncoder for nominal variables
encoder = OneHotEncoder(sparse_output=False)
for col in categoricalColumns:
    encoded_df = pd.DataFrame(
        encoder.fit_transform(dataFrame[[col]]),
        columns=[f"{col}_Dummy_{cat}" for cat in encoder.categories_[0]]
        )
        # Add encoded columns to the original DataFrame
    dataFrame = pd.concat([dataFrame, encoded_df], axis=1)
print(dataFrame.columns)
print(dataFrame)
#////////////////////////////////////////////////////////////////////////////////////////////

#3) We use Ordinal Encoding for only ordinal categorical variables.
# Instantiate the OrdinalEncoder
encoder = OrdinalEncoder()

for col in ordinalColumns:
    dataFrame[f"{col}_Numeric_Ordinal_Encoding"] = encoder.fit_transform(dataFrame[[col]]) + 1  # Add 1 to start levels from 1
print(dataFrame.columns)
print(dataFrame)
#********************************************************************************************<
#********************************************************************************************>
# We calculate and plot correlartion matrix for all numeric variables in our Final data set, including dummy variables
# Extract numeric columns
numeric_col = dataFrame.select_dtypes(exclude=['object']).columns

# Calculate the correlation matrix, ignoring null values
correlation_matrix = dataFrame[numeric_col].corr(method='pearson', min_periods=1)

# Plot the heatmap
plt.figure(figsize=(10, 8))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)

# Add title and labels
plt.title("Correlation Matrix Heatmap for all Numeric Features, using Final Data Set", fontsize=16)
plt.xlabel("Features", fontsize=12)
plt.ylabel("Features", fontsize=12)

# Display the heatmap
plt.tight_layout()  # Adjust layout to fit everything
plt.show()
#********************************************************************************************<

#********************************************************************************************<
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#********************************************************************************************>
# # We save the final data set in excel file for verification
# file_name = 'C:/Users/reema/OneDrive/Desktop/Machine Learning/Group Project/output.xlsx'
# sheet_name = 'NewSheet'

# try:
#     # Load the existing Excel file
#     workbook = load_workbook(file_name)
    
#     # Check if the sheet already exists
#     if sheet_name in workbook.sheetnames:
#         # Generate a unique name by appending a number
#         count = 1
#         while f"{sheet_name}_{count}" in workbook.sheetnames:
#             count += 1
#         sheet_name = f"{sheet_name}_{count}"
    
#     # Add the sheet to the existing file
#     with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
#         dataFrame.to_excel(writer, sheet_name=sheet_name, index=False)
#     print(f"Sheet '{sheet_name}' added to '{file_name}'.")
    
# except FileNotFoundError:
#     # Create a new file and add the sheet if it doesn't exist
#     with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
#         dataFrame.to_excel(writer, sheet_name=sheet_name, index=False)
#     print(f"Created new file '{file_name}' with sheet '{sheet_name}'.")


# #********************************************************************************************<
# #<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

#<><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><>

#********************************************************************************************>
#Logistic Redgression Model

# We define the selected features and target variable into new dataframe named logisticDataFrame
selected_features = [
    'transactionAmount', 
    'interestRate', 
    'risk', 
    'customerAgeGroup_Numeric_Ordinal_Encoding',  # Ordinal encoding
    'productCategory_Dummy_Electronics', 
    'productCategory_Dummy_Food', 
    'productCategory_Dummy_Furniture',
    'paymentMethod_Dummy_Credit Card', 
    'paymentMethod_Dummy_Debit Card']

# Create a new dataframe for Logistic Regression
logisticDataFrame = dataFrame[selected_features + ['class']]  # Include the target variable 'class'

# Display the first few rows of the new dataframe
print("The Data set will be used in Bulding logistic regression is as following:",'\n')
print(logisticDataFrame,'\n')
#-------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------
#Now we test the assumptions of Logistc Redredssion Model
#********************************************************************************************>
# Step 1: Check Multicollinearity
#********************************************************************************************>

# Define the initial feature set (before adding higher-order terms or interaction terms)
initial_features = ['transactionAmount', 'interestRate', 'risk']

# Add constant for VIF calculation
X_vif = add_constant(logisticDataFrame[initial_features])

# Calculate VIF for each feature
vif_data = pd.DataFrame()
vif_data["Variable"] = X_vif.columns
vif_data["VIF"] = [variance_inflation_factor(X_vif.values, i) for i in range(X_vif.shape[1])]
print("Variance Inflation Factor (VIF):\n", vif_data)

# Interpretation:
# - If VIF > 10, the variable is highly collinear and may need to be dropped or transformed.
# - Proceed to add higher-order terms and interaction terms only after addressing multicollinearity.
if any(vif_data["VIF"] > 10):
    for col, vif in zip(vif_data["Variable"], vif_data["VIF"]):
        if vif > 10:
            print(f"Feature '{col}' has VIF = {vif:.2f}. Consider removing it due to multicollinearity.")
else:
    print("No features need to be removed due to multicollinearity. All VIF values are less than 10.")

#********************************************************************************************>
# Step 2: Assess Linearity of the Logit
#********************************************************************************************>

# Define continuous variables for the linearity test
continuous_vars = ['transactionAmount', 'interestRate', 'risk']

# Add constant for Logistic Regression
X_cont = add_constant(logisticDataFrame[continuous_vars])
y = logisticDataFrame['class']

# Fit a logistic regression model
logit_model = Logit(y, X_cont).fit(disp=False)

# Calculate logit values (log-odds)
logisticDataFrame.loc[:, 'logit'] = np.log(logit_model.predict() / (1 - logit_model.predict()))


# Plot each continuous variable against the logit
fig, axes = plt.subplots(1, len(continuous_vars), figsize=(15, 5))
for i, var in enumerate(continuous_vars):
    axes[i].scatter(logisticDataFrame[var], logisticDataFrame['logit'], alpha=0.6)
    axes[i].set_title(f'{var} vs Logit')
    axes[i].set_xlabel(var)
    axes[i].set_ylabel('Logit')

plt.tight_layout()
plt.show()

# Interpretation:
# - If the scatter plots show a linear trend, the assumption of linearity for the logit is satisfied.
# - If the trend is not linear, consider transformations such as polynomial terms or log transformations 
#   for the continuous variables to meet the assumption.

#==========================================================================================
# Define continuous variables
continuous_vars_riskSQ = ['transactionAmount', 'interestRate', 'risk', 'risk_squared']

logisticDataFrameSQ = logisticDataFrame.copy()

# Transform risk variables
logisticDataFrameSQ.loc[:, 'risk_squared'] = logisticDataFrameSQ['risk'] ** 2

# Add constant for logistic regression
X_cont = add_constant(logisticDataFrameSQ[continuous_vars_riskSQ])
y = logisticDataFrameSQ['class']

# Fit the logistic regression model
logit_model = Logit(y, X_cont).fit(disp=False)

# # Calculate logit values
# logisticDataFrameSQ.loc[:, 'logit'] = np.log(logit_model.predict() / (1 - logit_model.predict()))


# Plot each continuous variable against logit
fig, axes = plt.subplots(1, len(continuous_vars_riskSQ), figsize=(15, 5))

for i, var in enumerate(continuous_vars_riskSQ):
    # Debugging step to check data validity
    if logisticDataFrameSQ[var].isnull().sum() > 0:
        print(f"Skipping {var} due to missing values.")
        continue

    # Plot scatter
    axes[i].scatter(logisticDataFrameSQ[var], logisticDataFrameSQ['logit'], alpha=0.6)
    axes[i].set_title(f'{var} vs Logit')
    axes[i].set_xlabel(var)
    axes[i].set_ylabel('Logit')

plt.tight_layout()
plt.show()


#********************************************************************************************>
# Step 3: Add Higher-Order and Interaction Terms
#********************************************************************************************>

# Create higher-order terms for 'risk' to capture non-linear relationships
logisticDataFrame.loc[:, 'risk_squared'] = logisticDataFrame['risk'] ** 2  # Squared term for 'risk'

# Create an interaction term between 'risk' and 'interestRate' to capture potential interaction effects
logisticDataFrame.loc[:, 'risk_interestRate_interaction'] = logisticDataFrame['risk'] * logisticDataFrame['interestRate']

#logisticDataFrame.drop (columns= 'risk', inplace = True)

# Display the modified DataFrame with the new terms added
print("\nLogistic Regression DataFrame after adding higher-order and interaction terms for 'risk':\n")
print(logisticDataFrame[['risk_squared', 'risk_interestRate_interaction']].head())

#********************************************************************************************>
# Step 4: Logistic Regression Model with New Terms
#********************************************************************************************>

# Define features for Logistic Regression, including new terms
features_without_Risk = [
    'transactionAmount',
    'interestRate',
    'risk_squared',  # Higher-order term for 'risk'
    'risk_interestRate_interaction',  # Interaction term for 'risk'
    'customerAgeGroup_Numeric_Ordinal_Encoding',  # Ordinal encoding
    'productCategory_Dummy_Electronics', 
    'productCategory_Dummy_Food', 
    'productCategory_Dummy_Furniture',
    'paymentMethod_Dummy_Credit Card', 
    'paymentMethod_Dummy_Debit Card']



# Define the feature matrix (X) and target variable (y)
X = add_constant(logisticDataFrame[features_without_Risk])  # Add a constant term
y = logisticDataFrame['class']  # Target variable

# Fit the logistic regression model with the new features
logit_model = Logit(y, X).fit()

# Display the summary of the logistic regression model
print("\nLogistic Regression Model Summary with New Terms:\n")
print(logit_model.summary())


# Data Preparation
# Define selected features
final_selected_features = [
    'transactionAmount', 
    'interestRate', 
    'risk', 
    'customerAgeGroup_Numeric_Ordinal_Encoding', 
    'productCategory_Dummy_Electronics', 
    'productCategory_Dummy_Food', 
    'productCategory_Dummy_Furniture',
    'paymentMethod_Dummy_Credit Card', 
    'paymentMethod_Dummy_Debit Card'
]

# Feature matrix (X) and target variable (y)
X = dataFrame[final_selected_features]
y = dataFrame['class']

# Split the dataset into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42, stratify=y)

# Oversample the minority class using SMOTE
smote = SMOTE(random_state=42)
X_train_balanced, y_train_balanced = smote.fit_resample(X_train, y_train)

# Polynomial Features
poly = PolynomialFeatures(degree=2, interaction_only=True, include_bias=False)
X_train_poly = poly.fit_transform(X_train_balanced)
X_test_poly = poly.transform(X_test)

# Hyperparameter Tuning for Logistic Regression
param_grid = {
    'C': [0.01, 0.1, 1, 10, 100],  # Regularization strength
    'penalty': ['l1', 'l2'],  # Type of regularization
    'solver': ['liblinear', 'saga']  # Compatible solvers for l1 and l2
}

logistic_model = LogisticRegression(class_weight='balanced', random_state=42, max_iter=1000)

grid_search = GridSearchCV(estimator=logistic_model, param_grid=param_grid, scoring='roc_auc', cv=5, verbose=1, n_jobs=-1)
grid_search.fit(X_train_poly, y_train_balanced)

# Get the best model
best_model = grid_search.best_estimator_
print("Best hyperparameters:", grid_search.best_params_)

# Predict on the test set
y_pred_default = best_model.predict(X_test_poly)
y_proba = best_model.predict_proba(X_test_poly)[:, 1]

# Evaluate Model Performance
print("Confusion Matrix (Default Threshold):")
print(confusion_matrix(y_test, y_pred_default))
print("\nClassification Report (Default Threshold):")
print(classification_report(y_test, y_pred_default))
print("ROC-AUC Score:", roc_auc_score(y_test, y_proba))

# Precision-Recall Curve and Threshold Adjustment
precision, recall, thresholds = precision_recall_curve(y_test, y_proba)

# Plot Precision-Recall Curve
plt.figure(figsize=(8, 6))
plt.plot(recall, precision, label="Precision-Recall Curve")
plt.xlabel("Recall")
plt.ylabel("Precision")
plt.title("Precision-Recall Curve")
plt.legend()
plt.grid()
plt.show()

# Plot ROC Curve
RocCurveDisplay.from_estimator(best_model, X_test_poly, y_test)
plt.title("ROC Curve")
plt.show()

# Adjust Decision Threshold
optimal_threshold = 0.4  # Adjust based on the precision-recall curve
y_pred_adjusted = (y_proba > optimal_threshold).astype(int)

print(f"\nConfusion Matrix (Threshold = {optimal_threshold}):")
print(confusion_matrix(y_test, y_pred_adjusted))
print(f"\nClassification Report (Threshold = {optimal_threshold}):")
print(classification_report(y_test, y_pred_adjusted))


#===========================================================================================

print("=================================================================================\n")
print("Model Evaluation Summary:")
print("- At the default threshold (0.5):")
print("  * Overall Accuracy: 53%")
print("  * Class 0 (Majority): Precision = 71%, Recall = 55%")
print("  * Class 1 (Minority): Precision = 31%, Recall = 47%")
print("  * ROC-AUC Score: 0.53 (slightly above random guessing)")

print("\n- At the adjusted threshold (0.4):")
print("  * Overall Accuracy: 42%")
print("  * Class 0: Precision = 75%, Recall = 25%")
print("  * Class 1: Precision = 31%, Recall = 80%")

print("\n- Insights:")
print("  * The model struggles to balance precision and recall due to class imbalance.")
print("  * Recall for the minority class improves at the cost of overall accuracy.")
print("  * Consider techniques to handle class imbalance or more advanced models to improve results.")


print("=================================================================================\n")

#===========================================================================================
